"""
NICE BizLine 산업경쟁현황 재무 적재 (비상장 포함 화장품사 매출·영업이익·광고비).

요청사항/NICE_BizLINE_산업경쟁현황_통합_tag.xlsx →
  - RawData: 기업 × 연도(2023~25) × 분류(매출액·영업이익·광고비) → nice_financials
  - 고유기업_태그: 기업 → 브랜드 텍스트 → nice_company_brands (우리 21개 브랜드 매칭)

연 단위 데이터 → 주간 지표와 분리된 '재무' 탭 전용. DART 대체(상장·비상장 통합).
수동/연 1회 적재: python -m signals.nice_financials [xlsx경로]
"""

import os
import logging

import openpyxl
from sqlalchemy import text

from config.settings import DB_SCHEMA
from config.brands import ALL_BRANDS, BRAND_KO_NAMES
from storage.models import get_session

logger = logging.getLogger(__name__)

_DEFAULT_XLSX = os.path.join(os.path.dirname(__file__), "..", "..",
                            "요청사항", "NICE_BizLINE_산업경쟁현황_통합_tag.xlsx")


def _ensure_tables(session) -> None:
    session.execute(text(f"""
        CREATE TABLE IF NOT EXISTS {DB_SCHEMA}.nice_financials (
            id BIGSERIAL PRIMARY KEY,
            industry_code VARCHAR(20),
            industry_name VARCHAR(120),
            company VARCHAR(200) NOT NULL,
            metric VARCHAR(20) NOT NULL,          -- 매출액 | 영업이익 | 광고비
            year INT NOT NULL,
            amount BIGINT,
            rank INT,
            is_cosmetic BOOLEAN DEFAULT FALSE,
            loaded_at TIMESTAMP WITH TIME ZONE DEFAULT NOW(),
            UNIQUE(industry_code, company, metric, year)
        )
    """))
    session.execute(text(f"""
        CREATE TABLE IF NOT EXISTS {DB_SCHEMA}.nice_company_brands (
            industry_code VARCHAR(20),
            company VARCHAR(200) NOT NULL,
            industry_name VARCHAR(120),
            is_cosmetic BOOLEAN DEFAULT FALSE,
            brand_tags TEXT,                       -- 원본 브랜드 텍스트
            matched_brands TEXT,                   -- 우리 모니터링 브랜드(쉼표)
            is_single_brand BOOLEAN DEFAULT FALSE, -- 회사≈단일브랜드(회사전체=브랜드단독 근사)
            PRIMARY KEY (industry_code, company)
        )
    """))
    session.execute(text(
        f"CREATE INDEX IF NOT EXISTS ix_nice_fin_company "
        f"ON {DB_SCHEMA}.nice_financials (company, metric, year)"))
    session.commit()


# 우리 21개 브랜드 → 한글 별칭(매칭용). BRAND_KO_NAMES + 영문.
def _brand_aliases() -> dict:
    d = {}
    for b in ALL_BRANDS:
        aliases = list(BRAND_KO_NAMES.get(b, []))
        aliases.append(b)                      # 영문명도
        d[b] = [a for a in aliases if a]
    return d


# 태그에 브랜드명이 없지만 운영사명으로 매칭되는 케이스(수동 보강)
_BRAND_COMPANY = {
    "Abib": ["포컴퍼니"], "Mixsoon": ["파켓"], "By Wishtrend": ["위시컴퍼니"],
}


def _match_monitored(brand_tags: str, company: str, aliases: dict) -> list:
    """브랜드 태그 텍스트(또는 운영사명)에서 우리 모니터링 브랜드 매칭."""
    t = brand_tags or ""
    co = company or ""
    hit = []
    for brand, al in aliases.items():
        if t and any(a and a in t for a in al):
            hit.append(brand)
        elif any(c in co for c in _BRAND_COMPANY.get(brand, [])):
            hit.append(brand)
    return hit


def load(xlsx_path: str = None) -> dict:
    """엑셀 → DB 적재. 반환 {financials, companies, mapped}."""
    path = xlsx_path or _DEFAULT_XLSX
    if not os.path.exists(path):
        logger.error("NICE 엑셀 없음: %s", path)
        return {"financials": 0, "companies": 0, "mapped": 0}

    aliases = _brand_aliases()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    session = get_session()
    fin_n = comp_n = mapped_n = 0
    try:
        _ensure_tables(session)
        # 재적재 전 클리어(연 1회 전량 갱신)
        session.execute(text(f"TRUNCATE {DB_SCHEMA}.nice_financials, {DB_SCHEMA}.nice_company_brands"))

        # ── RawData (배치 INSERT — 원격 DB 성능) ──
        ws = wb["RawData"]
        fin_sql = text(f"""
            INSERT INTO {DB_SCHEMA}.nice_financials
                (industry_code, industry_name, company, metric, year, amount, rank, is_cosmetic)
            VALUES (:ic,:inm,:co,:m,:y,:amt,:rk,:cos)
        """)
        batch, seen_fin = [], set()
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or r[0] is None or r[3] is None:
                continue
            ind_code, ind_name, metric, company, year, amount, rank, cos = r[:8]
            try:
                year = int(year); amount = int(amount) if amount is not None else None
                rank = int(rank) if rank is not None else None
            except (TypeError, ValueError):
                continue
            key = (ind_code, company, metric, year)   # UNIQUE 제약 — 소스 중복 제거
            if key in seen_fin:
                continue
            seen_fin.add(key)
            batch.append({"ic": ind_code, "inm": ind_name, "co": company, "m": metric,
                          "y": year, "amt": amount, "rk": rank, "cos": (str(cos).upper() == "Y")})
            if len(batch) >= 1000:
                session.execute(fin_sql, batch); fin_n += len(batch); batch = []
        if batch:
            session.execute(fin_sql, batch); fin_n += len(batch)
        session.commit()

        # ── 고유기업_태그 (배치, 중복키 dedup) ──
        ws2 = wb["고유기업_태그"]
        cb_sql = text(f"""
            INSERT INTO {DB_SCHEMA}.nice_company_brands
                (industry_code, company, industry_name, is_cosmetic, brand_tags, matched_brands, is_single_brand)
            VALUES (:ic,:co,:inm,:cos,:bt,:mb,:sb)
        """)
        cbatch, seen = [], set()
        for r in ws2.iter_rows(min_row=2, values_only=True):
            if not r or r[0] is None or r[2] is None:
                continue
            ind_code, ind_name, company, cos = r[0], r[1], r[2], r[3]
            key = (ind_code, company)
            if key in seen:
                continue
            seen.add(key)
            brand_tags = str(r[6]) if len(r) > 6 and r[6] else ""
            matched = _match_monitored(brand_tags, company, aliases)
            single = bool(brand_tags) and ("," not in brand_tags) and ("OEM" not in brand_tags)
            cbatch.append({"ic": ind_code, "co": company, "inm": ind_name,
                           "cos": (str(cos).upper() == "Y"), "bt": brand_tags,
                           "mb": ",".join(matched), "sb": single})
            comp_n += 1
            if matched:
                mapped_n += 1
            if len(cbatch) >= 1000:
                session.execute(cb_sql, cbatch); cbatch = []
        if cbatch:
            session.execute(cb_sql, cbatch)
        session.commit()
    finally:
        session.close()
        wb.close()
    logger.info("NICE 적재: 재무 %d행 · 기업 %d · 우리브랜드매칭 %d", fin_n, comp_n, mapped_n)
    return {"financials": fin_n, "companies": comp_n, "mapped": mapped_n}


if __name__ == "__main__":
    import sys
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    print(load(sys.argv[1] if len(sys.argv) > 1 else None))
